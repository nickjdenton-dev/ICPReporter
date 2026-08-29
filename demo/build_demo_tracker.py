#!/usr/bin/env python3
"""Build a fictional QC Trend Tracker workbook for tutorial videos.

All Data uses the same headers Spec. Reporter Pro commits into, so a new demo
run can be appended. Product tabs carry Excel line charts (measured / claim /
warn / limit) with ~16 historical points each.
"""
from __future__ import annotations

import math
import random
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parent / "QC-Trend-Tracker-DEMO.xlsx"

N_POINTS = 16
WARN_PCT = 0.05
LIMIT_PCT = 0.10
# Last history date is before the demo raw-data run (2026-08-28) so a commit
# lands as a new point on the right of every chart.
START = date(2026, 3, 12)
STEP_DAYS = 11

COMPOUND = {
    "P": "Phosphorus (P₂O₅)",
    "K": "Potassium (K₂O)",
    "Ca": "Calcium (Ca)",
    "Mg": "Magnesium (Mg)",
    "B": "Boron (B)",
    "Zn": "Zinc (Zn)",
    "Cu": "Copper (Cu)",
    "Mn": "Manganese (Mn)",
    "Fe": "Iron (Fe)",
}
WL = {
    "P": "213.618",
    "K": "766.491",
    "Ca": "393.366",
    "Mg": "279.553",
    "B": "249.772",
    "Zn": "213.857",
    "Cu": "324.754",
    "Mn": "257.610",
    "Fe": "238.204",
}

# (name, [(el, claim), ...], extras)
# extras: which point indexes dip below warn / limit for a realistic QC story
PRODUCTS = [
    ("FieldGreen Org", [("P", 3.0), ("K", 3.0), ("Ca", 4.0), ("Cu", 0.1), ("Zn", 0.1)],
     {"warn": (5,), "limit": (12,)}),
    ("Root Humic", [("Ca", 0.5), ("Fe", 0.5), ("Mg", 0.4), ("Mn", 0.15), ("Zn", 0.05), ("B", 0.02)],
     {"warn": (3,), "limit": ()}),
    ("Root Humic with Metals", [("Ca", 0.5), ("Fe", 0.5), ("Mg", 0.4), ("Mn", 0.15), ("Zn", 0.05), ("B", 0.02), ("Cu", 0.01)],
     {"warn": (8,), "limit": ()}),
    ("Zinc Slurry", [("Zn", 1.0)],
     {"warn": (6,), "limit": (14,)}),
    ("Humic Filler", [("Ca", 0.4), ("Fe", 0.1), ("Mg", 0.05)],
     {"warn": (4,), "limit": ()}),
    ("MicroMix-8", [("B", 3.0), ("Mn", 3.5), ("Zn", 4.0), ("Fe", 1.0), ("Cu", 0.5), ("Mg", 0.5)],
     {"warn": (7,), "limit": (11,)}),
    ("K-Sol", [("K", 20.0)],
     {"warn": (2,), "limit": (13,)}),
    ("CalPhos Gold", [("P", 14.0), ("Ca", 10.0)],
     {"warn": (9,), "limit": ()}),
    ("GreenLeaf 7-7-7", [("P", 7.0), ("K", 7.0)],
     {"warn": (1,), "limit": (10,)}),
    ("TraceRx", [("Mg", 1.0), ("B", 0.02), ("Cu", 0.1), ("Fe", 2.0), ("Mn", 2.0), ("Zn", 0.5)],
     {"warn": (6,), "limit": ()}),
]

HEADERS = [
    "Date", "SampleType", "Batch", "Element", "Measured", "LabelClaim", "Flag",
    "SourceFile", "ImportedAt", "Submitted By", "Table", "Limit (%)", "Detected",
    "Wavelength (nm)", "Instrument",
]
NAVY = "0F4C81"
GREEN = "2E9B48"
ORANGE = "E08A1E"
RED = "C0392B"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
THIN = Border(
    left=Side(style="thin", color="C2C8CF"),
    right=Side(style="thin", color="C2C8CF"),
    top=Side(style="thin", color="C2C8CF"),
    bottom=Side(style="thin", color="C2C8CF"),
)


def dates():
    return [START + timedelta(days=STEP_DAYS * i) for i in range(N_POINTS)]


def walk(claim: float, rng: random.Random, warn_idx, limit_idx) -> list[float]:
    """AR(1) wander around the claim, with optional warn/limit dips."""
    x = claim * rng.uniform(0.99, 1.03)
    out = []
    sigma = max(claim * 0.012, 0.0004)
    for i in range(N_POINTS):
        x = 0.65 * x + 0.35 * claim + rng.gauss(0, sigma)
        seasonal = claim * 0.008 * math.sin(i / 3.2)
        v = max(1e-6, x + seasonal)
        out.append(v)
    for i in warn_idx:
        out[i] = claim * (1 - WARN_PCT) * rng.uniform(0.985, 0.998)
    for i in limit_idx:
        out[i] = claim * (1 - LIMIT_PCT) * rng.uniform(0.96, 0.995)
    return [round(v, 6) for v in out]


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN


def color_series(chart):
    colors = [NAVY, GREEN, ORANGE, RED]
    for i, s in enumerate(chart.series):
        s.graphicalProperties.line.solidFill = colors[i % 4]
        s.graphicalProperties.line.width = 18000  # EMUs
        s.marker = Marker(symbol=None)
        s.smooth = False


def add_line_chart(ws, title, min_col, min_row, max_row, anchor):
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = "% w/w"
    chart.y_axis.axId = 100
    chart.x_axis.numFmt = "YYYY-MM-DD"
    chart.x_axis.majorGridlines = None
    chart.legend.position = "b"
    chart.height = 7.2
    chart.width = 14
    data = Reference(ws, min_col=min_col + 1, min_row=min_row, max_col=min_col + 4, max_row=max_row)
    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    color_series(chart)
    ws.add_chart(chart, anchor)
    return chart


def safe_sheet(name: str, taken: set) -> str:
    base = "".join(ch if ch not in r':\/?*[]' else " " for ch in name).strip()[:31]
    if not base:
        base = "Sheet"
    out = base
    n = 2
    while out.lower() in taken:
        suffix = f" {n}"
        out = (base[: 31 - len(suffix)] + suffix)
        n += 1
    taken.add(out.lower())
    return out


def build():
    rng = random.Random(41012)
    dts = dates()
    wb = Workbook()

    # --- cover ---
    cover = wb.active
    cover.title = "Read me"
    cover["A1"] = "Spec. Reporter Pro — DEMO QC tracker"
    cover["A1"].font = Font(name="Calibri", size=18, bold=True, color=NAVY)
    cover.merge_cells("A1:F1")
    cover["A3"] = (
        "Fictional history for tutorial videos. 16 committed lots per sample type "
        "(March–August 2026), so each chart already has a trend. The matching instrument "
        "export is demo/raw-data-qc-demo.xlsx (dated 28 Aug 2026) — commit that run onto "
        "this workbook to show a new point landing on an existing chart."
    )
    cover["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    cover.merge_cells("A3:F6")
    cover.row_dimensions[3].height = 72
    cover["A8"] = "How to use in the video"
    cover["A8"].font = Font(bold=True, color=NAVY, size=13)
    steps = [
        "1. Import demo/sample-types-demo.json (Edit → Sample Types → Import) if those types are not already in the bank.",
        "2. In Spec. Reporter Pro (Chrome/Edge): Tracker → Select tracker file… → pick this workbook.",
        "3. Drop demo/raw-data-qc-demo.xlsx, lock/finalize, then commit the new lots.",
        "4. Charts are on each product tab. All Data is the sheet the app appends to — do not rename it.",
        "5. Warning band is 5% below claim (orange); limit is 10% below claim (red). Green is the claim.",
    ]
    for i, s in enumerate(steps):
        cover[f"A{9+i}"] = s
        cover[f"A{9+i}"].alignment = Alignment(wrap_text=True)
        cover.row_dimensions[9+i].height = 22
    cover.column_dimensions["A"].width = 110
    cover.sheet_properties.tabColor = "0F4C81"

    # --- All Data ---
    ad = wb.create_sheet("All Data", 1)
    for c, h in enumerate(HEADERS, 1):
        ad.cell(1, c, h)
    style_header(ad, len(HEADERS))
    ad.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    ad.freeze_panes = "A2"
    widths = {
        "A": 12, "B": 26, "C": 10, "D": 22, "E": 12, "F": 12, "G": 8,
        "H": 28, "I": 20, "J": 16, "K": 12, "L": 12, "M": 12, "N": 16, "O": 24,
    }
    for col, w in widths.items():
        ad.column_dimensions[col].width = w

    all_rows = []  # dicts in header order
    per_product = {}  # name -> {el: [(date, measured, claim), ...]}

    for p_i, (pname, analytes, extras) in enumerate(PRODUCTS):
        per_product[pname] = {}
        warn_idx = extras.get("warn") or ()
        limit_idx = extras.get("limit") or ()
        for el, claim in analytes:
            vals = walk(claim, rng, warn_idx, limit_idx)
            series = []
            for j, (d, meas) in enumerate(zip(dts, vals)):
                batch = str(40000 + p_i * 100 + j + 1)
                imported = f"{d.isoformat()} 14:22:00"
                table = "Macro" if el in ("P", "K") else "Micro"
                all_rows.append({
                    "Date": d.isoformat(),
                    "SampleType": pname,
                    "Batch": batch,
                    "Element": COMPOUND[el],
                    "Measured": meas,
                    "LabelClaim": claim,
                    "Flag": "",
                    "SourceFile": "Spec. Reporter Pro v5.18.5",
                    "ImportedAt": imported,
                    "Submitted By": "Demo Analyst",
                    "Table": table,
                    "Limit (%)": "",
                    "Detected": "",
                    "Wavelength (nm)": WL[el],
                    "Instrument": "Agilent ICP-OES 5800",
                })
                series.append((d, meas, claim))
            per_product[pname][el] = series

    all_rows.sort(key=lambda r: (r["Date"], r["SampleType"], r["Element"]))
    for r_i, rec in enumerate(all_rows, 2):
        for c, h in enumerate(HEADERS, 1):
            val = rec[h]
            cell = ad.cell(r_i, c, val)
            cell.border = THIN
            cell.font = Font(name="Calibri", size=10)
            if h == "Date":
                cell.number_format = "YYYY-MM-DD"
            if h in ("Measured", "LabelClaim") and isinstance(val, float):
                cell.number_format = "0.0000"
    ad.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(all_rows)+1}"

    # --- product chart tabs ---
    taken = {"read me", "all data"}
    for pname, analytes, _extras in PRODUCTS:
        sheet_name = safe_sheet(pname, taken)
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = "2E9B48"
        ws["A1"] = f"{pname} — QC trend charts"
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
        ws.merge_cells("A1:E1")
        ws["A2"] = "Charts update when Spec. Reporter Pro commits data. Orange = 5% below claim; red = 10% below claim."
        ws["A2"].font = Font(italic=True, color="666666", size=10)
        ws.merge_cells("A2:E2")

        # One analyte block after another: data in A–E, chart to the right at G.
        for a_i, (el, claim) in enumerate(analytes):
            header_row = 4 + a_i * (N_POINTS + 4)
            series = per_product[pname][el]
            headers = [f"{COMPOUND[el]} date", "Measured", "Claim", "Warn", "Limit"]
            for j, h in enumerate(headers):
                cell = ws.cell(header_row, 1 + j, h)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
            for k, (d, meas, cl) in enumerate(series):
                rr = header_row + 1 + k
                ws.cell(rr, 1, d).number_format = "YYYY-MM-DD"
                ws.cell(rr, 2, meas).number_format = "0.0000"
                ws.cell(rr, 3, cl).number_format = "0.0000"
                ws.cell(rr, 4, round(cl * (1 - WARN_PCT), 6)).number_format = "0.0000"
                ws.cell(rr, 5, round(cl * (1 - LIMIT_PCT), 6)).number_format = "0.0000"
            last_row = header_row + N_POINTS
            add_line_chart(ws, COMPOUND[el], 1, header_row, last_row, f"G{header_row}")
        for col, w in zip("ABCDE", (20, 12, 12, 12, 12)):
            ws.column_dimensions[col].width = w
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["A"].width = 22
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_view.showGridLines = False

    # hide the cover? keep visible so they see instructions first
    wb.save(OUT)
    print(f"wrote {OUT} rows={len(all_rows)} products={len(PRODUCTS)} points={N_POINTS}")


if __name__ == "__main__":
    build()
